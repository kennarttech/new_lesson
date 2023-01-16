"""this reverse a list========="""
numbers  = [22, 14, 20, 1, 43]

print(numbers[::])
print(len(numbers))


if (len(numbers)) > 5:
    print('the numbers are more than 5')
else:
    print('the numbers are less 5')




secret_number = 10
guess_count = 0
guess_limit = 3

while guess_count < guess_limit:
    """===============Guessing number==========="""
    guess = int(input('Please Guess a number: '))
    guess_count += 1
    if guess == secret_number:
        print('You won!')
        break
else:
    print('Your guess is incorrect ):')



command = ""
started = False
options = ['(start = start the car)\n(stop = stop the car)' ]


for m in options:
    print(m)

while True:
    command = input('Enter ur option...').lower()
    if command == 'start':
        if started:
            print('Car is already started')
        else:
            started = True
            print('Car started.....')
    elif command == 'stop':
        if not started:
            print('Car already stopped')
        else:
            started = False
            print('Car stopped.')
        break
    else:
        print("Sorry i don't understand the option")
        print(options)





prices =  [10, 20, 30]
total = 0

for price in prices:
    total += price
print(f'Total: {price}')




for x in range(4):
    for y in range(3):
        print(f'{x}, {y}')





numbers = [5, 2, 5, 2, 2]

for x in numbers:
    output = ''
    for count in range(x):
        output += 'x'
    print(output)




numbers = [2,3,4,5,3]
new_numbers = []

for number in numbers:
    if number not in new_numbers:
        new_numbers.append(number)
print(new_numbers)




message = input('>')
"""Dictionary with for loop"""
words = message.split(' ')

emojis = {":)": "😔😁", ":(": "😔"}
output = ""

for word in words:
    output += emojis.get(word, word) + " "
    print(output)




def greet_user(age, username):
    """This function take user input and return the value"""
    return age + username

output = input('Please enter your name: ')


result = greet_user(str(3), username=output)
print(result)





def greatness(first, last):
    """function using keyword agmt b4 pos agmt"""
    return f'{first} {last}'


result = greatness(first='kenneth', last='lumor')
print(result)





def sum(a, b):
  return a + b

result = sum(3, 4)
print(result)  # prints 7




def factorial(n):
    if n == 0:
        return 1
    else:
        return n * factorial(n - 1)
result = int(input('Please enter number: '))

print(factorial(result))  # Output: 120




# =========================function=========
def emoji_converter(message):
    words = message.split(" ")
    emojis = {":": "😁", ":": "😔"}
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


message = input(">: ")
result = emoji_converter(message)
print(result)





try:
    """Exceptiong handling"""
    age = int(input('Please enter your age: '))
    income = 2000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print('Age cannot be 0')
except ValueError:
    print('Invalid value')





import openpyxl as xl 
from openpyxl.chart import BarChart, Reference

file_path = '/home/kenneth/Desktop/New_python_class/variables/transactions.xlsx'

wb = xl.load_workbook(filename=file_path) 
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)


for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price


values = Reference(sheet, min_row=2, 
        max_row=sheet.max_row,
        min_col=4,
        max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')
wb.save('transaction2.xlsx')





file_path = '/home/kenneth/Desktop/New_python_class/variables/transactions.xlsx'


def process_workbook(filename):
    wb = xl.load_workbook(filename=file_path) 
    sheet = wb['Sheet1']


    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price


    values = Reference(sheet, min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)

result = process_workbook(filename=file_path)






class Student:
    def __init__(self, name, age, grade):
        self.name = name
        self.age = age
        self.grade = grade

    def get_name(self):
        return self.name
    
    def get_age(self):
        return self.age
    
    def get_grade(self):
        return self.grade


student1 = Student("Alice", 16, "A")
student2 = Student("Bob", 17, "B")


name1 = student1.get_name()
age2 = student2.get_age()
grade1 = student1.grade


print(name1)  
print(age2)  
print(grade1)





class Car:
    def __init__(self, make, model, year):
        self.make = make
        self.model = model
        self.year = year
        self.odometer_reading = 0

    def get_description(self):
        long_name = f"{self.year} {self.make} {self.model}"
        return long_name.title()

    def read_odometer(self):
        print(f"This car has {self.odometer_reading} miles on it.")

    def update_odometer(self, mileage):
        if mileage >= self.odometer_reading:
            self.odometer_reading = mileage
        else:
            print("You can't roll back an odometer!")

    def increment_odometer(self, miles):
        self.odometer_reading += miles



my_car = Car('audi', 'a4', 2016)
print(my_car.get_description())
my_car.update_odometer(23500)
my_car.read_odometer()
my_car.increment_odometer(100)
my_car.read_odometer()


# 2016 Audi A4
# This car has 23500 miles on it.
# This car has 23600 miles on it.





class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y
        
    def distance_from_origin(self):
        return (self.x ** 2 + self.y ** 2) ** 0.5
    
    def __str__(self):
        return f"({self.x}, {self.y})"



point1 = Point(1, 2)
point2 = Point(3, 4)

print(point1.distance_from_origin()) 
print(point2.distance_from_origin()) 

print(point1)  
print(point2) 


